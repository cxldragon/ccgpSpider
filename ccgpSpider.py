#-*- coding: UTF-8 -*-

#+++++++++++++++++++++++++++++++++++++++++++++
#Created By CxlDragon 2018.8.30
#用于抓取中国政府采购网（ccgp）的招标公告信息
#+++++++++++++++++++++++++++++++++++++++++++++

#Updted 2018.8.31
#把url中的搜索时间类型做为参数可配置
#修复了一次连续爬取多页时最后一页序列号错误等bug
#启用all_page参数，会爬取结果的总页数，限定起始页和终止页
#在配置文件中保存最后一条记录的时间，用于控制只爬最新记录

#Updated 2018.8.30
#增加了json文件，所有配置从config.json文件中读取

#Updated 2018.8.29
#增加邮件发送功能，把查询的结果以邮件形式发至指定邮箱

import sys
import time
import urllib
import requests                  #pip3 install requests
import numpy as np               #pip3 install numpy
from bs4 import BeautifulSoup    #pip3 install beautifulsoup4
from openpyxl import Workbook    #pip3 install openpyxl
from openpyxl import load_workbook
from imp import reload
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
import smtplib
#import os.path
import json
from dojson import JsonConf
from datetime import datetime


reload(sys)

#全局参数
all_page=0           #总页数
page_begin=1         #起始的开爬页
page_end=1           #结束页（即最大爬到第几页）
each_page_rows=20    #每页总记录条数
if_send_mail = True  #是否发送邮件
last_datetime = ""   #最新的记录时间，如果为空则爬所有数据，如果有值则根据最后时间判断新记录
search_keywords=""   #搜索用的关键字，多个关键字 用 + 连接即可
time_type = "1"      #搜索的时间类型：0.今日 1.近三天 2.近一周 3.近一月 4.近三月 5.近半年 6.指定日期
# User Agents 设置   {'User-Agent':''},\
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52'}]

# 邮件信息
mail_info = {
    "from": "xxxx@xxx.cn",
    "to": "xxxx@xxx.cn",
    "hostname": "smtp.xxx.cn",
    "username": "xxxx@xxx.cn",
    "password": "xxxx",
    "mail_subject": "",
    "mail_text": "",
    "mail_encoding": "utf-8",
    "mail_att":""
    }

#++++++++++++++++++++
#爬列表并解析
#page_begin:开爬页 page_end:结束页 zb_tag:爬关键字
#++++++++++++++++++++
def lists_spider(page_begin,page_end,zb_tag):
    #声明全局变量，否则为只读，一旦修改了则定义为局部变量
    global all_page     #页面数量
    global mail_info    #邮件信息
    global last_datetime #最后时间
    
    zb_list=[]
    zb_list_mail=[] #用于发送邮件
    try_times=0
    new_time=""  #datetime.now().strftime('%Y.%m.%d %H:%M:%S')
    out_count=0
    
    #out_count=page_begin*each_page_rows-each_page_rows
    
    while(1):

        #最近三天
        url='http://search.ccgp.gov.cn/bxsearch?searchtype=1&page_index='+str(page_begin)+'&bidSort=0&buyerName=&projectId=&pinMu=0&bidType=0&dbselect=bidx&kw='+urllib.request.quote(zb_tag)+'&start_time=&end_time=&timeType='+time_type+'&displayZone=&zoneId=&pppStatus=0&agentName='
        #print(url)
        time.sleep(np.random.rand()*5)  #随机时延，避免被禁ip        
        #采用这种方式爬，ip不容易被禁止，更换浏览代理
        try:
            req = urllib.request.Request(url, headers=hds[page_begin%len(hds)])
            source_code = urllib.request.urlopen(req).read()
            plain_text=str(source_code.decode('utf-8')) #指明用utf-8解码
        except (urllib.HTTPError, urllib.URLError) :
            print("somethings is error...")
            continue

        soup = BeautifulSoup(plain_text, 'lxml') #对获得的页面用bs4排序，用html.parser lxml 等解析器

        #如果是第一次，则取页面总数
        if all_page == 0:
            try:
                page_soup=soup.find('p',{'class':'pager'})      #获取包含页数的soup
                if page_soup!=None:
                    #处理java脚本
                  
                    pagerstr=page_soup.find('script').get_text()
                    pagerstr=pagerstr.split(',')[0].split(':')[1].strip()
                    all_page=int(pagerstr)
                    print('最多可爬页面总数 %d 页（配置的起始页第 %d 页，结束页是第 %d 页）' %(all_page,page_begin,page_end))
                else:
                    all_page=1
                    print('获取最多可爬页面总数失败page_soup')
            except:
                all_page=1
                print('获取最多可爬页面总数失败：默认为1，结束页也重置为1')
                #break  #取不到总页数退出，注释掉则表示取不到总页数，仍然执行，但只会取一页即begin_page
            
        #确保所取的页在总页数范围在(1…all_page)之间
        #注意：在此page_begin和page_end 不是全局变量，因此只在函数内有效
        if page_begin<=0: page_begin=1
        if page_begin>all_page:page_begin=all_page
        if page_end<=0: page_end=1
        if page_end>all_page:page_end=all_page
        if out_count == 0:
            out_count=page_begin*each_page_rows-each_page_rows
            print('实际爬页面从第 %d 页开始，爬到第 %d 页结束' %(page_begin,page_end))
    
        #获取查询结果列表
        try:
            #获取查询结果列表
            list_soup = soup.find('ul',{'class': 'vT-srch-result-list-bid'})
        except:
            print('解析查询结果列表失败')
            list_soup = None

        # 连续5次取不到request的信息，则退出
        try_times+=1;
        if list_soup==None and try_times<5:
            continue
        elif list_soup==None or len(list_soup)<=1:
            print('给出的url取不到需要的内容')
            break
        print('开始解析第 %d 页的列表…' % page_begin)
        #导出excel的文件名
        save_path='zb_list-'+zb_tag+'.xlsx'
        #开始循环处理网页
        licount=0
        for zb_info in list_soup.findAll('li'):
            title = str(zb_info.find('a',{'style':'line-height:18px'}).get_text()).strip()
            
            zb_url = zb_info.find('a', {'style':'line-height:18px'}).get('href')            
            content = str(zb_info.find('p').get_text()).strip()
            desc = str(zb_info.find('span').get_text()).strip()
            desc_list = desc.split('|')
            
            try:
                time_info = '' + str('|'.join(desc_list[0:1])).strip()  #时间：
                #取最大的时间 new_time为空或者小于当前记录的时间，则赋值
                if new_time=="" or (datetime.strptime(new_time,'%Y.%m.%d %H:%M:%S')<datetime.strptime(time_info,'%Y.%m.%d %H:%M:%S')): 
                    new_time=time_info
            except:
                time_info ='时间：暂无'
            try:
                zbcg_info = str(desc_list[1]).strip().split('：')[1]   #采购人
            except:
                zbcg_info = '采购人：暂无'
            try:
                zbdl_info_list =  str(desc_list[2]).split("\r\n")   #代理机构 + 公告类型  .replace(" ",'')
                zbdl_info = zbdl_info_list[0].strip().split('：')[1]   #代理机构
                zb_type = zbdl_info_list[-4].strip()    #公告类型
                #zb_type=desc_list[2]
            except:
                zbdl_info = '代理：暂无'
                zb_type = '公告类型：暂无'
            try:
                zb_city=str(desc_list[3]).strip()   #招标区域城市
            except:
                zb_city='城市：暂无'
            try:
                zbcg_bd=str(desc_list[4]).strip()   #采购标的
            except:
                zbcg_bd='标的：暂无'

            #判断是否是只取最新
            if last_datetime !="":
                try:
                    before_time=datetime.strptime(last_datetime,'%Y.%m.%d %H:%M:%S')  #上次爬完后保存的时间
                    #print("上次保存时间：%s 第%d 页第 %d 条记录时间：%s" % (before_time,page_begin,out_count+1,time_info))
                    if before_time>=datetime.strptime(time_info,'%Y.%m.%d %H:%M:%S'): #只导出最新记录                        
                        continue
                except:
                    last_datetime=""

            #print("时间=%s 标题=%s 区域=%s 公告类型=%s 标的=%s zbcg=%s zbdl=%s url=%s"
            #      %(time_info,title,zb_city,zb_type,zbcg_bd,zbcg_info,zbdl_info,zb_url))

            #获取url的文本
            zb_all=zb_content(zb_url,page_begin%len(hds))
            zb_list.append([time_info,title,zb_city,zb_type,zbcg_bd,zbcg_info,zbdl_info,zb_url,content,zb_all])
            #是否发送邮件
            if if_send_mail:
                zb_list_mail.append([time_info,zb_city,zb_type,zbcg_info,title,zb_url])
            licount+=1    #本页输出计数
            out_count+=1  #总输出计数            
            try_times=0 #只要成功获取一次数据，则把尝试次数复位为 0

        #判断是否有需要导出的内容
        print("爬到第 %d 页的记录 %d 条" % (page_begin,licount))
        if licount>=1:
            #print("成功爬到第 %d 页的记录 %d 条" % (page_begin,licount))
            if page_begin==1:
                zb_list2excel(zb_list,zb_tag,True,out_count-licount,save_path)
            else:
                zb_list2excel(zb_list,zb_tag,False,out_count-licount,save_path)            
        
        #清空list，准备爬下一页
        page_begin+=1
        del zb_list[:]
        #是否爬到了最后一页
        if page_begin>page_end:
            #发送邮件标识为“真”且爬到的新记录
            if if_send_mail and out_count>(page_begin-1)*each_page_rows-each_page_rows:
                #发送邮件
                mail_info["mail_subject"]="ccgp公告-"+zb_tag+"-"+time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
                mail_text = "<html><body></p>"
                ic=1
                for bl in zb_list_mail:
                    mail_text = mail_text+"【"+str(ic)+"】\t"+bl[0]+"\t"+bl[1]+"\t"+bl[2]+"\t"+bl[3]+"\t"+bl[4]+"\t</br><a href="+bl[5]+">"+bl[5]+"</a><p/>"
                    ic+=1
                mail_text+="</body></html>"
                #添加正文信息
                mail_info["mail_text"]=mail_text
                #添加附件信息
                mail_info["mail_att"]=save_path
                #执行发送邮件
                send_mail(mail_info)

            last_datetime=new_time  #更新最后的记录时间
            print("最新一条记录的时间为："+last_datetime)
            break  #退出while(1)循环
    #循环结束过程返回
    return


#++++++++++++++++++++
#把爬到的招标lits信息导出到excel表格中
#bfirst:是否第一次导出 begin_count导出的开始记录序号
#save_path:导出保存的文件名
#++++++++++++++++++++
def zb_list2excel(zb_list,zb_tag,bfirst,begin_count,save_path):
    try:
        if bfirst:  #是否第一次写入文件
            wb=Workbook()
            ws = wb.active
            ws.title=zb_tag
            ws.append(['序号','时间','标题','区域','公告类型','标的','采购方','招标代理','url','内容','全文'])
        else:
            wb=load_workbook(save_path)
            ws = wb.active
            ws.title=zb_tag

        count=1
        for bl in zb_list:
            ws.append([begin_count+count,bl[0],bl[1],bl[2],bl[3],bl[4],bl[5],bl[6],bl[7],bl[8],bl[9]])
            count+=1
        wb.save(save_path)
        print('导出excel文件成功！本次导出 %d 条，累计导出记录 %d 条' %(count-1,begin_count+count-1))
    except:
        print('执行导出excel文件出错！')
    return


#++++++++++++++++++++
#爬公告的详情页面
#useragent_index:指定模拟的浏览器
#++++++++++++++++++++
def zb_content(url,useragent_index):
    result = ''
    
    for i in range(0,3): # 连续3次取不到request的信息，则退出
        time.sleep(np.random.rand()*5)  #随机时延，避免被禁ip        
        #采用这种方式爬，ip不容易被禁止，更换浏览代理
        try:
            req = urllib.request.Request(url, headers=hds[useragent_index])
            source_code = urllib.request.urlopen(req).read()
            plain_text=str(source_code.decode('utf-8')) #指明用utf-8解码
        except (urllib.HTTPError, urllib.URLError) :
            print("somethings is error...")
            continue
  
        soup = BeautifulSoup(plain_text, 'html.parser')  #对获得的页面用bs4排序，用html解析器
        main_soup = soup.find('div',{'class': ['vF_detail_main','vT_detail_main']})

        if main_soup==None and i<3: #三次取不到就退出
            continue
        elif main_soup==None or len(main_soup)<=1 or i==4:
            print('给出的url取不到需要的内容')
            return result
        else:   #成功取取数据
            break

    try:
        content_soup = main_soup.find('div',{'class':['vF_detail_content','vT_detail_content w760c']})  #vF_detail_content vT_detail_content w760c
        if content_soup!=None:
            result=str(content_soup.get_text()).strip()
        else:
            result=''
            print('解析url获取的内容为空 %s' %(url))
    except:
        print('解析url获取的内容出错 %s' %(url))
        result=''
    
    #print(result)
    return result

#++++++++++++++++++++
#通过ssl发送邮件
#++++++++++++++++++++
def send_mail(mail_info):
    #这里使用SMTP_SSL就是默认使用465端口
    try:
        smtp = smtplib.SMTP_SSL(mail_info["hostname"])
        print('登录邮箱…')
        #smtp.set_debuglevel(1)#设置调试日志级别
        
        smtp.ehlo(mail_info["hostname"])
        smtp.login(mail_info["username"], mail_info["password"])
        print('登录成功')
        #邮件内容初始化
        msg = MIMEMultipart()
        #邮件正文
        #msg.attach(MIMEText(mail_info["mail_text"], "plain", mail_info["mail_encoding"]))  #text文本格式
        msg.attach(MIMEText(mail_info["mail_text"], "html", mail_info["mail_encoding"]))   #html格式
        msg["Subject"] = Header(mail_info["mail_subject"], mail_info["mail_encoding"])
        msg["from"] = mail_info["from"]
        msg["to"] = mail_info["to"]
        #添加附件
        if mail_info["mail_att"]!="" :
            sendfile=open(mail_info["mail_att"],'rb').read()
            text_att = MIMEText(sendfile, 'base64', 'utf-8') 
            text_att["Content-Type"] = 'application/octet-stream'
            text_att.add_header('Content-Disposition', 'attachment', filename=mail_info["mail_att"])
            msg.attach(text_att)

        print("邮件开始发送…")
        smtp.sendmail(mail_info["from"], mail_info["to"], msg.as_string())
        smtp.quit()
        print("邮件发送成功")
    except smtplib.SMTPException as e:
        print("邮件发送失败",e)
    return
        
#++++++++++++++++++++
#加载配置文件
#++++++++++++++++++++
def load_cfg():
    global page_begin     #起始的开爬页
    global page_end       #结束页
    global each_page_rows #每页记录数
    global if_send_mail   #是否发邮件
    global mail_info      #邮件信息
    global hds            #http agent
    global last_datetime  #最后时间
    global search_keywords #搜索关键字
    global time_type      #搜索的时间类型

    cfg=JsonConf.load()
    try:
        #先加载常量，防止丢失数据
        hds=cfg["hds"]
        mail_info=cfg["mail_info"]
        each_page_rows=int(cfg["each_page_rows"])
        if_send_mail=bool(cfg["if_send_mail"])
        #再加载变量
        page_begin=int(cfg["page_begin"])
        page_end=int(cfg["page_end"])
        last_datetime=cfg["last_datetime"]
        search_keywords=cfg["search_keywords"]
        time_type=cfg["time_type"]
        print("成功加载配置文件")
    except:
        print("配置文件加载失败")    
    return
    
#++++++++++++++++++++
#保存配置文件
#++++++++++++++++++++
def save_cfg():
    global mail_info
    cfg={}
    cfg["page_begin"]=str(page_begin)
    cfg["page_end"]=str(page_end)
    cfg["each_page_rows"]=str(each_page_rows)
    cfg["if_send_mail"]=bool(if_send_mail)
    #邮件主题，内容，附件不保存
    mail_info["mail_subject"]=""
    mail_info["mail_text"]=""
    mail_info["mail_att"]=""
    cfg["mail_info"]=mail_info
    cfg["hds"]=hds
    cfg["last_datetime"]=last_datetime
    cfg["search_keywords"]=search_keywords
    cfg["time_type"]=time_type
    JsonConf.save(cfg)
    return

if __name__=='__main__':

    #运行参数配置
    #加载配置
    load_cfg()

    #爬的关键字 多个关键字，用 + 号连接
    if search_keywords!="":
        zb_tag=search_keywords
    else:
        zb_tag="河长"

    #开爬
    zb_list=lists_spider(page_begin,page_end,zb_tag)

    #保存配置
    save_cfg()


